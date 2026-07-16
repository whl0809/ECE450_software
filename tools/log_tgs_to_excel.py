#!/usr/bin/env python3
"""Log six TGS sensor channels from ADS114S06 and export an Excel workbook.

Recommended repository path:
    ECE450_software/tools/log_tgs_to_excel.py

Examples:
    # Hardware logging for 60 seconds, 1 sample per second
    sudo env GPIOZERO_PIN_FACTORY=lgpio python3 tools/log_tgs_to_excel.py \
        --duration 60 \
        --interval 1 \
        --output data/tgs_logs/tgs_test.xlsx

    # Test without hardware
    python3 tools/log_tgs_to_excel.py \
        --mock \
        --duration 10 \
        --interval 1 \
        --output data/tgs_logs/mock_tgs_test.xlsx

    # Log and also update the display JSON for the CO5300 TGS dashboard
    sudo env GPIOZERO_PIN_FACTORY=lgpio python3 tools/log_tgs_to_excel.py \
        --duration 120 \
        --interval 1 \
        --write-json runtime/tgs_state.json
"""

from __future__ import annotations

import argparse
import json
import math
import random
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Missing openpyxl. Install it with:\n"
        "  sudo apt update\n"
        "  sudo apt install -y python3-openpyxl"
    ) from exc


CHANNELS: list[tuple[str, int]] = [
    ("TGS2610", 0),
    ("TGS2620", 1),
    ("TGS2603", 2),
    ("TGS2602", 3),
    ("TGS2600", 4),
    ("TGS2611", 5),
]

SIGNAL_NAMES: list[str] = [
    "TGS2610_VOUT",
    "TGS2620_VOUT",
    "TGS2603_VOUT",
    "TGS2602_VOUT",
    "TGS2600_VOUT",
    "TGS2611_VOUT",
]


# ADS114S06 commands/registers.
# This script uses the same simple diagnostic setup as the previous terminal monitor.
CMD_RESET = 0x06
CMD_START = 0x08
CMD_RDATA = 0x12
CMD_RREG = 0x20
CMD_WREG = 0x40

REG_ID = 0x00
REG_INPMUX = 0x02
REG_PGA = 0x03
REG_DATARATE = 0x04
REG_REF = 0x05

AINCOM = 0x0C


@dataclass
class ChannelReading:
    name: str
    ain: int
    raw: int | None
    voltage_v: float | None
    ready: bool
    error: str = ""


@dataclass
class Sample:
    sample_index: int
    wall_time: str
    elapsed_s: float
    status: str
    readings: list[ChannelReading]


def signed16(msb: int, lsb: int) -> int:
    return int.from_bytes(bytes([msb, lsb]), byteorder="big", signed=True)


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    tmp.replace(path)


class MockAdsReader:
    def __init__(self, vref: float) -> None:
        self.vref = vref
        self.step = 0

    def open(self) -> None:
        pass

    def configure(self) -> int:
        return 0x05

    def close(self) -> None:
        pass

    def read_channel(self, ain: int) -> tuple[int, float, bool]:
        base = [0.82, 1.05, 1.34, 1.55, 0.96, 1.18][ain]
        voltage = base + 0.16 * math.sin(self.step / 4 + ain * 0.7) + random.uniform(-0.01, 0.01)
        voltage = max(0.0, min(self.vref, voltage))
        raw = int(voltage / self.vref * 32768)
        if ain == 5:
            self.step += 1
        return raw, voltage, True


class ADS114S06:
    def __init__(
        self,
        *,
        bus: int,
        device: int,
        speed: int,
        mode: int,
        vref: float,
        gain: float,
        data_rate_code: int,
        use_gpio: bool,
        start_gpio: int,
        drdy_gpio: int,
        timeout: float,
        settle: float,
    ) -> None:
        try:
            import spidev
        except ImportError as exc:
            raise SystemExit(
                "Missing spidev. Install it with:\n"
                "  sudo apt update\n"
                "  sudo apt install -y python3-spidev"
            ) from exc

        self.spidev = spidev
        self.bus = bus
        self.device = device
        self.speed = speed
        self.mode = mode
        self.vref = vref
        self.gain = gain
        self.data_rate_code = data_rate_code
        self.timeout = timeout
        self.settle = settle
        self.spi = spidev.SpiDev()
        self.start_pin = None
        self.drdy_pin = None

        if use_gpio:
            try:
                from gpiozero import DigitalInputDevice, OutputDevice
            except ImportError as exc:
                raise SystemExit(
                    "Missing gpiozero/lgpio. Install it with:\n"
                    "  sudo apt update\n"
                    "  sudo apt install -y python3-gpiozero python3-lgpio\n"
                    "Run with:\n"
                    "  sudo env GPIOZERO_PIN_FACTORY=lgpio python3 tools/log_tgs_to_excel.py ..."
                ) from exc

            if start_gpio >= 0:
                # START is kept low. Conversions are started by command 0x08.
                self.start_pin = OutputDevice(start_gpio, active_high=True, initial_value=False)

            if drdy_gpio >= 0:
                # DRDY is active low.
                self.drdy_pin = DigitalInputDevice(drdy_gpio, pull_up=False)

    def open(self) -> None:
        self.spi.open(self.bus, self.device)
        self.spi.max_speed_hz = self.speed
        self.spi.mode = self.mode
        self.spi.bits_per_word = 8

    def close(self) -> None:
        try:
            self.spi.close()
        except Exception:
            pass
        if self.start_pin is not None:
            self.start_pin.close()
        if self.drdy_pin is not None:
            self.drdy_pin.close()

    def xfer(self, data: list[int]) -> list[int]:
        return self.spi.xfer2(data)

    def reset(self) -> None:
        self.xfer([CMD_RESET])
        time.sleep(0.005)

    def read_register(self, reg: int) -> int:
        rx = self.xfer([CMD_RREG | (reg & 0x1F), 0x00, 0x00])
        return rx[-1]

    def write_register(self, reg: int, value: int) -> None:
        self.xfer([CMD_WREG | (reg & 0x1F), 0x00, value & 0xFF])

    def configure(self) -> int:
        self.reset()
        device_id = self.read_register(REG_ID)

        # Provisional diagnostic setup:
        # PGA gain x1, project default data-rate code, external reference path.
        self.write_register(REG_PGA, 0x00)
        self.write_register(REG_DATARATE, self.data_rate_code & 0xFF)
        self.write_register(REG_REF, 0x00)

        return device_id

    def select_channel(self, ain: int) -> None:
        # Positive input = AINx, negative input = AINCOM.
        mux = ((ain & 0x0F) << 4) | AINCOM
        self.write_register(REG_INPMUX, mux)

    def wait_ready(self) -> bool:
        if self.drdy_pin is None:
            time.sleep(self.settle)
            return True

        time.sleep(self.settle)
        deadline = time.monotonic() + self.timeout
        while time.monotonic() < deadline:
            if self.drdy_pin.value == 0:
                return True
            time.sleep(0.001)
        return False

    def read_channel(self, ain: int) -> tuple[int, float, bool]:
        self.select_channel(ain)
        self.xfer([CMD_START])
        ready = self.wait_ready()
        rx = self.xfer([CMD_RDATA, 0x00, 0x00])
        raw = signed16(rx[-2], rx[-1])
        voltage = (raw / 32768.0) * (self.vref / self.gain)
        return raw, voltage, ready


def create_reader(args: argparse.Namespace) -> Any:
    if args.mock:
        return MockAdsReader(args.vref)

    return ADS114S06(
        bus=args.spi_bus,
        device=args.spi_device,
        speed=args.speed,
        mode=args.mode,
        vref=args.vref,
        gain=args.gain,
        data_rate_code=args.data_rate_code,
        use_gpio=not args.no_gpio,
        start_gpio=args.start_gpio,
        drdy_gpio=args.drdy_gpio,
        timeout=args.timeout,
        settle=args.settle,
    )


def read_one_sample(reader: Any, sample_index: int, t0: float) -> Sample:
    readings: list[ChannelReading] = []
    status = "OK"

    for name, ain in CHANNELS:
        try:
            raw, voltage, ready = reader.read_channel(ain)
            if not ready:
                status = "DRDY_TIMEOUT"
            readings.append(ChannelReading(name, ain, raw, voltage, ready))
        except Exception as exc:
            status = "READ_ERROR"
            readings.append(ChannelReading(name, ain, None, None, False, str(exc)))

    return Sample(
        sample_index=sample_index,
        wall_time=datetime.now().isoformat(timespec="seconds"),
        elapsed_s=time.monotonic() - t0,
        status=status,
        readings=readings,
    )


def print_sample(sample: Sample) -> None:
    fields = []
    for reading in sample.readings:
        if reading.voltage_v is None:
            fields.append(f"{reading.name}=ERR")
        else:
            fields.append(f"{reading.name}={reading.voltage_v:.4f}V")
    print(
        f"[{sample.sample_index:05d}] "
        f"t={sample.elapsed_s:8.2f}s "
        f"status={sample.status:<12} "
        + "  ".join(fields),
        flush=True,
    )


def sample_to_dashboard_payload(sample: Sample) -> dict[str, Any]:
    return {
        "tgs": [
            {
                "name": reading.name,
                "raw": reading.raw,
                "voltage_v": None if reading.voltage_v is None else round(reading.voltage_v, 6),
            }
            for reading in sample.readings
        ],
        "sample_hz": None,
        "system_status": sample.status,
        "updated_at": sample.wall_time,
    }


def build_workbook(samples: list[Sample], args: argparse.Namespace) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Samples"

    headers = ["sample_index", "wall_time", "elapsed_s", "status"]
    for name, _ain in CHANNELS:
        headers.extend([f"{name}_raw", f"{name}_voltage_v", f"{name}_ready"])
    ws.append(headers)

    for sample in samples:
        row: list[Any] = [
            sample.sample_index,
            sample.wall_time,
            round(sample.elapsed_s, 3),
            sample.status,
        ]
        for reading in sample.readings:
            row.extend([
                reading.raw,
                None if reading.voltage_v is None else round(reading.voltage_v, 6),
                reading.ready,
            ])
        ws.append(row)

    # Style Samples sheet.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 14,
        "B": 22,
        "C": 12,
        "D": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col_idx in range(5, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"

    # Summary sheet.
    summary = wb.create_sheet("Summary")
    summary.append(["Field", "Value"])
    summary.append(["Generated at", datetime.now().isoformat(timespec="seconds")])
    summary.append(["Output mode", "mock" if args.mock else "hardware"])
    summary.append(["Sample count", len(samples)])

    if samples:
        duration = samples[-1].elapsed_s - samples[0].elapsed_s
        hz = (len(samples) - 1) / duration if duration > 0 and len(samples) > 1 else None
        summary.append(["First wall time", samples[0].wall_time])
        summary.append(["Last wall time", samples[-1].wall_time])
        summary.append(["Duration seconds", round(duration, 3)])
        summary.append(["Estimated sample Hz", None if hz is None else round(hz, 3)])
    else:
        summary.append(["First wall time", ""])
        summary.append(["Last wall time", ""])
        summary.append(["Duration seconds", 0])
        summary.append(["Estimated sample Hz", ""])

    summary.append([])
    summary.append(["Sensor", "AIN", "avg_raw", "min_voltage_v", "avg_voltage_v", "max_voltage_v", "valid_samples"])

    for channel_index, (name, ain) in enumerate(CHANNELS):
        raws: list[int] = []
        voltages: list[float] = []
        for sample in samples:
            reading = sample.readings[channel_index]
            if reading.raw is not None:
                raws.append(reading.raw)
            if reading.voltage_v is not None:
                voltages.append(reading.voltage_v)

        summary.append([
            name,
            ain,
            round(sum(raws) / len(raws), 2) if raws else None,
            round(min(voltages), 6) if voltages else None,
            round(sum(voltages) / len(voltages), 6) if voltages else None,
            round(max(voltages), 6) if voltages else None,
            len(voltages),
        ])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for cell in summary[11]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 22
    for col_idx in range(3, 8):
        summary.column_dimensions[get_column_letter(col_idx)].width = 18

    for row in summary.iter_rows(min_row=2, min_col=1, max_col=summary.max_column):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"

    notes = wb.create_sheet("Notes")
    notes.append(["Note"])
    notes.append(["This workbook contains raw ADC codes and VOUT voltages only."])
    notes.append(["Do not interpret these values as ppm until per-sensor calibration is complete."])
    notes.append(["TGS mapping: AIN0=TGS2610, AIN1=TGS2620, AIN2=TGS2603, AIN3=TGS2602, AIN4=TGS2600, AIN5=TGS2611."])
    notes.column_dimensions["A"].width = 110
    notes["A1"].fill = header_fill
    notes["A1"].font = header_font

    return wb


def default_output_path() -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path("data") / "tgs_logs" / f"tgs_log_{stamp}.xlsx"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Read six TGS channels and export an Excel workbook.")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path.")
    parser.add_argument("--duration", type=float, default=60.0, help="Logging duration in seconds.")
    parser.add_argument("--interval", type=float, default=1.0, help="Seconds between samples.")
    parser.add_argument("--write-json", type=Path, default=None, help="Optionally update runtime/tgs_state.json for display.")
    parser.add_argument("--mock", action="store_true", help="Generate simulated data without hardware.")
    parser.add_argument("--no-gpio", action="store_true", help="Do not use START/DRDY GPIO; sleep instead.")

    parser.add_argument("--spi-bus", type=int, default=0)
    parser.add_argument("--spi-device", type=int, default=0)
    parser.add_argument("--speed", type=int, default=1_000_000)
    parser.add_argument("--mode", type=int, default=1)
    parser.add_argument("--vref", type=float, default=4.096)
    parser.add_argument("--gain", type=float, default=1.0)
    parser.add_argument("--data-rate-code", type=lambda x: int(x, 0), default=0x14)

    parser.add_argument("--start-gpio", type=int, default=17)
    parser.add_argument("--drdy-gpio", type=int, default=27)
    parser.add_argument("--timeout", type=float, default=0.5)
    parser.add_argument("--settle", type=float, default=0.05)

    return parser


def main() -> int:
    args = build_parser().parse_args()

    if args.duration <= 0:
        raise SystemExit("--duration must be positive")
    if args.interval <= 0:
        raise SystemExit("--interval must be positive")

    output = args.output if args.output is not None else default_output_path()
    output.parent.mkdir(parents=True, exist_ok=True)

    reader = create_reader(args)
    samples: list[Sample] = []
    t0 = time.monotonic()

    print("ECE450 TGS Excel Logger")
    print(f"mode       : {'mock' if args.mock else 'hardware'}")
    print(f"duration   : {args.duration:.1f} s")
    print(f"interval   : {args.interval:.3f} s")
    print(f"output     : {output}")
    if args.write_json:
        print(f"display js : {args.write_json}")
    print("Press Ctrl+C to stop early and still save Excel.")
    print()

    try:
        reader.open()
        device_id = reader.configure()
        if not args.mock:
            print(f"ADS114S06 ID register: 0x{device_id:02X}")

        sample_index = 0
        next_time = time.monotonic()
        end_time = t0 + args.duration

        while time.monotonic() < end_time:
            now = time.monotonic()
            if now < next_time:
                time.sleep(min(0.05, next_time - now))
                continue

            sample = read_one_sample(reader, sample_index, t0)
            samples.append(sample)
            print_sample(sample)

            if args.write_json:
                atomic_write_json(args.write_json, sample_to_dashboard_payload(sample))

            sample_index += 1
            next_time += args.interval

    except KeyboardInterrupt:
        print("\nStopped early by user.")
    finally:
        reader.close()

    wb = build_workbook(samples, args)
    wb.save(output)

    print()
    print(f"[PASS] Saved Excel workbook: {output}")
    print(f"       Samples: {len(samples)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
